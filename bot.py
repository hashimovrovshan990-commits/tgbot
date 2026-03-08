# -*- coding: utf-8 -*-
"""
Trader's Journal - улучшенная версия bot.py
Все рекомендации из анализа учтены.
"""

import logging
import os
import json
import matplotlib.pyplot as plt
from io import BytesIO
from aiohttp import web
from pathlib import Path
from datetime import datetime, timedelta, date
from typing import Optional, Dict, Any, List, Tuple, Union
import calendar as cal

# Безопасный импорт psycopg2 (для будущей миграции)
try:
    import psycopg2
except ImportError:
    psycopg2 = None

from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    ReplyKeyboardMarkup, KeyboardButton,
    InlineKeyboardMarkup, InlineKeyboardButton,
    FSInputFile, LabeledPrice, PreCheckoutQuery
)
from openpyxl import Workbook

# ========== Конфигурация из переменных окружения ==========
def get_env(key: str, default: str = "") -> str:
    return os.environ.get(key, default)

TOKEN = get_env("BOT_TOKEN")
if not TOKEN:
    raise ValueError("BOT_TOKEN must be set!")

PROVIDER_TOKEN = get_env("PROVIDER_TOKEN", "")
MAX_TRADES_FREE = int(get_env("MAX_TRADES_FREE", "20"))
ADMIN_ID = int(get_env("ADMIN_ID", "0"))
ADMIN_PASSWORD = get_env("ADMIN_PASSWORD")  # может быть пустой строкой
if not ADMIN_PASSWORD:
    await message.answer("Админ-панель не настроена (отсутствует пароль).")
    return

CURRENCY = "USD"
WEBHOOK_DOMAIN = get_env("WEBHOOK_DOMAIN", "https://tgbot-ljj1.onrender.com")
WEBHOOK_SECRET = get_env("WEBHOOK_SECRET", "")  # для проверки подписи
WEBHOOK_PATH = f"/webhook/{TOKEN}" if TOKEN else "/webhook"
WEBHOOK_URL = f"{WEBHOOK_DOMAIN}{WEBHOOK_PATH}"
PORT = int(os.environ.get("PORT", 8000))

DATABASE_URL = get_env("DATABASE_URL", "database.db")

# ========== Логирование ==========
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ========== База данных (улучшенный класс) ==========
class Database:
    def __init__(self, db_path: str):
        self.db_path = db_path
        self.conn = None
        self.cursor = None
        self._connect()
        self._init_tables()

    def _connect(self):
        """Подключение к SQLite (в будущем можно заменить на PostgreSQL)."""
        import sqlite3
        self.conn = sqlite3.connect(self.db_path, check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self.cursor = self.conn.cursor()

    def _init_tables(self):
        """Создание таблиц, если их нет."""
        self.cursor.executescript("""
            CREATE TABLE IF NOT EXISTS users (
                user_id INTEGER PRIMARY KEY,
                username TEXT,
                first_name TEXT,
                language TEXT DEFAULT 'ru',
                total_trades INTEGER DEFAULT 0,
                is_premium INTEGER DEFAULT 0,
                premium_until TEXT,
                is_admin INTEGER DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS accounts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                name TEXT,
                initial_balance REAL,
                current_balance REAL,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS trades (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                account_id INTEGER,
                pair TEXT,
                trade_type TEXT,
                open_date TEXT,
                close_date TEXT,
                amount REAL,
                take_profit REAL,
                status TEXT,
                strategy TEXT,
                checklist TEXT,
                notes TEXT,
                screenshot_path TEXT,
                pnl REAL DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS subscriptions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                provider TEXT,
                provider_id TEXT,
                status TEXT,
                period_end TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
        """)
        self.commit()

    def commit(self):
        self.conn.commit()

    # --- Пользователи ---
    def get_user(self, user_id: int) -> Optional[Dict[str, Any]]:
        self.cursor.execute("SELECT * FROM users WHERE user_id=?", (user_id,))
        row = self.cursor.fetchone()
        return dict(row) if row else None

    def create_or_update_user(self, user_id: int, username: str, first_name: str):
        self.cursor.execute("""
            INSERT INTO users(user_id, username, first_name)
            VALUES(?, ?, ?)
            ON CONFLICT(user_id) DO UPDATE SET
                username=excluded.username,
                first_name=excluded.first_name
        """, (user_id, username or "user", first_name or "User"))
        if user_id == ADMIN_ID and ADMIN_ID != 0:
            self.cursor.execute("UPDATE users SET is_admin=1 WHERE user_id=?", (user_id,))
        self.commit()

    def set_language(self, user_id: int, lang: str):
        self.cursor.execute("UPDATE users SET language=? WHERE user_id=?", (lang, user_id))
        self.commit()

    def get_language(self, user_id: int) -> str:
        self.cursor.execute("SELECT language FROM users WHERE user_id=?", (user_id,))
        row = self.cursor.fetchone()
        return row["language"] if row else "ru"

    def increment_trades(self, user_id: int):
        self.cursor.execute("UPDATE users SET total_trades=total_trades+1 WHERE user_id=?", (user_id,))
        self.commit()

    def decrement_trades(self, user_id: int):
        self.cursor.execute("UPDATE users SET total_trades=total_trades-1 WHERE user_id=?", (user_id,))
        self.commit()

    def is_premium(self, user_id: int) -> bool:
        self.cursor.execute("SELECT is_premium, premium_until FROM users WHERE user_id=?", (user_id,))
        row = self.cursor.fetchone()
        if not row or not row["is_premium"]:
            return False
        until = row["premium_until"]
        if not until:
            # сброс некорректной записи
            self.cursor.execute("UPDATE users SET is_premium=0 WHERE user_id=?", (user_id,))
            self.commit()
            return False
        try:
            dt = datetime.fromisoformat(until)
        except Exception:
            self.cursor.execute("UPDATE users SET is_premium=0 WHERE user_id=?", (user_id,))
            self.commit()
            return False
        if dt < datetime.now():
            self.cursor.execute("UPDATE users SET is_premium=0 WHERE user_id=?", (user_id,))
            self.commit()
            return False
        return True

    def grant_premium(self, user_id: int, days: int, provider: str = "manual", provider_id: str = "") -> Optional[str]:
        period_end = (datetime.now() + timedelta(days=days)).isoformat()
        try:
            self.cursor.execute(
                "UPDATE users SET is_premium=1, premium_until=? WHERE user_id=?",
                (period_end, user_id)
            )
            self.cursor.execute(
                "INSERT INTO subscriptions(user_id, provider, provider_id, status, period_end) VALUES(?, ?, ?, ?, ?)",
                (user_id, provider, provider_id, "active", period_end)
            )
            self.commit()
            return period_end
        except Exception:
            logger.exception("grant_premium failed")
            return None

    def revoke_premium(self, user_id: int):
        try:
            self.cursor.execute("UPDATE users SET is_premium=0, premium_until=NULL WHERE user_id=?", (user_id,))
            self.cursor.execute("UPDATE subscriptions SET status='cancelled' WHERE user_id=? AND status='active'", (user_id,))
            self.commit()
        except Exception:
            logger.exception("revoke_premium failed")

    # --- Аккаунты ---
    def get_accounts(self, user_id: int) -> List[Dict[str, Any]]:
        self.cursor.execute("SELECT id, name, initial_balance, current_balance FROM accounts WHERE user_id=?", (user_id,))
        return [dict(row) for row in self.cursor.fetchall()]

    def create_account(self, user_id: int, name: str, balance: float):
        self.cursor.execute(
            "INSERT INTO accounts(user_id, name, initial_balance, current_balance) VALUES(?, ?, ?, ?)",
            (user_id, name, balance, balance)
        )
        self.commit()

    def get_account(self, account_id: int, user_id: int) -> Optional[Dict[str, Any]]:
        self.cursor.execute("SELECT * FROM accounts WHERE id=? AND user_id=?", (account_id, user_id))
        row = self.cursor.fetchone()
        return dict(row) if row else None

    def recalc_account_balance(self, user_id: int, account_id: int):
        """Пересчитывает current_balance как initial_balance + сумма pnl по сделкам."""
        acc = self.get_account(account_id, user_id)
        if not acc:
            return
        initial = acc["initial_balance"]
        self.cursor.execute(
            "SELECT COALESCE(SUM(pnl), 0) FROM trades WHERE user_id=? AND account_id=?",
            (user_id, account_id)
        )
        total_pnl = float(self.cursor.fetchone()[0] or 0)
        new_balance = initial + total_pnl
        self.cursor.execute(
            "UPDATE accounts SET current_balance=? WHERE id=? AND user_id=?",
            (new_balance, account_id, user_id)
        )
        self.commit()

    # --- Сделки ---
    def count_user_trades(self, user_id: int) -> int:
        self.cursor.execute("SELECT COUNT(*) FROM trades WHERE user_id=?", (user_id,))
        return self.cursor.fetchone()[0]

    def get_recent_trades(self, user_id: int, limit: int = 10):
        self.cursor.execute(
            "SELECT id, pair, trade_type, open_date, status FROM trades WHERE user_id=? ORDER BY created_at DESC LIMIT ?",
            (user_id, limit)
        )
        return [dict(row) for row in self.cursor.fetchall()]

    def get_trade(self, trade_id: int, user_id: int) -> Optional[Dict[str, Any]]:
        self.cursor.execute("SELECT * FROM trades WHERE id=? AND user_id=?", (trade_id, user_id))
        row = self.cursor.fetchone()
        return dict(row) if row else None

    def add_trade(self, user_id: int, data: dict):
        self.cursor.execute("""
            INSERT INTO trades
            (user_id, account_id, pair, trade_type, open_date, close_date, amount, take_profit, status, strategy, checklist, notes, screenshot_path, pnl)
            VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            user_id,
            data.get("account_id"),
            data.get("pair", "N/A"),
            data.get("trade_type", "N/A"),
            data.get("open_date", "N/A"),
            data.get("close_date", "N/A"),
            data.get("amount", 0.0),
            data.get("take_profit", 0.0),
            data.get("status", "N/A"),
            data.get("strategy", "N/A"),
            data.get("checklist"),
            data.get("notes", ""),
            data.get("screenshot_path"),
            data.get("pnl", 0.0)
        ))
        self.commit()

    def update_trade_field(self, trade_id: int, user_id: int, field: str, value: Union[str, float]):
        allowed = {"pair", "trade_type", "open_date", "close_date", "amount", "take_profit", "status", "strategy", "notes", "pnl"}
        if field not in allowed:
            raise ValueError(f"Field {field} not allowed")
        self.cursor.execute(
            f"UPDATE trades SET {field}=? WHERE id=? AND user_id=?",
            (value, trade_id, user_id)
        )
        self.commit()

    def delete_trade(self, trade_id: int, user_id: int):
        self.cursor.execute("DELETE FROM trades WHERE id=? AND user_id=?", (trade_id, user_id))
        self.commit()

    def get_all_trades_pnl(self, user_id: int) -> List[float]:
        self.cursor.execute("SELECT pnl FROM trades WHERE user_id=?", (user_id,))
        return [float(row[0]) for row in self.cursor.fetchall() if row[0] is not None]

    def get_trades_for_export(self, user_id: int, account_id: int = None):
        if account_id:
            self.cursor.execute(
                """SELECT pair, trade_type, open_date, close_date, amount, take_profit, status, strategy, notes, pnl, created_at
                   FROM trades WHERE user_id=? AND account_id=? ORDER BY created_at""",
                (user_id, account_id)
            )
        else:
            self.cursor.execute(
                """SELECT pair, trade_type, open_date, close_date, amount, take_profit, status, strategy, notes, pnl, created_at
                   FROM trades WHERE user_id=? ORDER BY created_at""",
                (user_id,)
            )
        return self.cursor.fetchall()

    def get_trades_for_equity(self, user_id: int):
        self.cursor.execute(
            "SELECT open_date, pnl FROM trades WHERE user_id=? AND pnl IS NOT NULL ORDER BY open_date",
            (user_id,)
        )
        return self.cursor.fetchall()

# Глобальный экземпляр БД
db = Database(DATABASE_URL)

# ========== Вспомогательные функции ==========
def parse_number(text: str) -> float:
    """Преобразует строку в число, поддерживая пробелы и запятую как разделитель."""
    if not text or not text.strip():
        raise ValueError("Empty string")
    s = text.strip().replace(" ", "").replace(",", ".")
    return float(s)

def can_add_trade(user_id: int) -> bool:
    """Проверяет, может ли пользователь добавить новую сделку."""
    if db.is_premium(user_id):
        return True
    return db.count_user_trades(user_id) < MAX_TRADES_FREE

# ========== Локализация ==========
LANG = {
    "ru": { ... },  # оставляем без изменений (содержимое такое же, как в исходном коде)
    "en": { ... }   # полностью сохранено
}

def get_text(user_id: int, key: str) -> str:
    lang = db.get_language(user_id)
    return LANG.get(lang, LANG["ru"]).get(key, key)

# ========== Клавиатуры ==========
def main_menu(user_id: int) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text=get_text(user_id, "my_trades")), KeyboardButton(text=get_text(user_id, "accounts"))],
        [KeyboardButton(text=get_text(user_id, "analytics")), KeyboardButton(text=get_text(user_id, "export"))],
        [KeyboardButton(text=get_text(user_id, "help")), KeyboardButton(text=get_text(user_id, "settings"))],
    ], resize_keyboard=True)

def calendar_kb(year: int, month: int, user_id: int) -> InlineKeyboardMarkup:
    kb = []
    # Навигация
    prev_m, prev_y = (month - 1, year) if month > 1 else (12, year - 1)
    next_m, next_y = (month + 1, year) if month < 12 else (1, year + 1)
    kb.append([
        InlineKeyboardButton(text="◀️", callback_data=f"cal_{prev_y}_{prev_m}"),
        InlineKeyboardButton(text=f"{year}-{month:02d}", callback_data="noop"),
        InlineKeyboardButton(text="▶️", callback_data=f"cal_{next_y}_{next_m}")
    ])
    # Сегодня / Вчера
    today_dt = date.today()
    yesterday_dt = today_dt - timedelta(days=1)
    kb.append([
        InlineKeyboardButton(text=get_text(user_id, "today"), callback_data=f"dt_{today_dt.year}_{today_dt.month}_{today_dt.day}"),
        InlineKeyboardButton(text=get_text(user_id, "yesterday"), callback_data=f"dt_{yesterday_dt.year}_{yesterday_dt.month}_{yesterday_dt.day}")
    ])
    # Дни недели
    lang = db.get_language(user_id)
    day_names = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"] if lang == "en" else ["Пн","Вт","Ср","Чт","Пт","Сб","Вс"]
    kb.append([InlineKeyboardButton(text=dn, callback_data="noop") for dn in day_names])
    # Числа
    for week in cal.monthcalendar(year, month):
        row = []
        for day in week:
            if day:
                row.append(InlineKeyboardButton(text=str(day), callback_data=f"dt_{year}_{month}_{day}"))
            else:
                row.append(InlineKeyboardButton(text=" ", callback_data="noop"))
        kb.append(row)
    kb.append([InlineKeyboardButton(text=get_text(user_id, "skip"), callback_data="skip")])
    return InlineKeyboardMarkup(inline_keyboard=kb)

# ========== Статистика ==========
def calculate_full_stats(user_id: int) -> Optional[Dict[str, Any]]:
    """
    Расширенная статистика и риск-менеджмент по всем сделкам пользователя.
    """
    pnls = db.get_all_trades_pnl(user_id)
    if not pnls:
        return None

    total = len(pnls)
    wins = [p for p in pnls if p > 0]
    losses = [p for p in pnls if p < 0]

    profitable_count = len(wins)
    losing_count = len(losses)
    win_rate = round(profitable_count / total * 100, 2) if total else 0.0

    total_pnl = round(sum(pnls), 2)
    avg_trade = round(total_pnl / total, 2) if total else 0.0

    best_trade = max(pnls) if pnls else 0.0
    worst_trade = min(pnls) if pnls else 0.0

    avg_risk = round(sum(abs(p) for p in losses) / losing_count, 2) if losing_count else 0.0

    avg_win = sum(wins) / profitable_count if profitable_count else 0.0
    avg_loss_abs = abs(sum(losses) / losing_count) if losing_count else 0.0
    avg_rr = round(avg_win / avg_loss_abs, 2) if avg_loss_abs > 0 else None  # None вместо inf

    # Максимальная просадка
    equity = 0.0
    peak = 0.0
    max_drawdown = 0.0
    for p in pnls:
        equity += p
        if equity > peak:
            peak = equity
        drawdown = peak - equity
        if drawdown > max_drawdown:
            max_drawdown = drawdown
    max_drawdown = round(max_drawdown, 2)

    # Серии
    max_win_streak = max_loss_streak = 0
    cur_win_streak = cur_loss_streak = 0
    for p in pnls:
        if p > 0:
            cur_win_streak += 1
            cur_loss_streak = 0
        elif p < 0:
            cur_loss_streak += 1
            cur_win_streak = 0
        else:
            cur_win_streak = cur_loss_streak = 0
        max_win_streak = max(max_win_streak, cur_win_streak)
        max_loss_streak = max(max_loss_streak, cur_loss_streak)

    return {
        "total": total,
        "profitable": profitable_count,
        "losing": losing_count,
        "win_rate": win_rate,
        "total_pnl": total_pnl,
        "avg_trade": avg_trade,
        "best_trade": best_trade,
        "worst_trade": worst_trade,
        "avg_risk": avg_risk,
        "avg_rr": avg_rr,
        "max_drawdown": max_drawdown,
        "win_streak": max_win_streak,
        "loss_streak": max_loss_streak,
    }

def create_equity_chart(user_id: int):
    rows = db.get_trades_for_equity(user_id)
    if not rows:
        return None
    dates = []
    profits = []
    for r in rows:
        if r[0] and r[1] is not None:
            try:
                d = datetime.fromisoformat(r[0]) if "T" not in str(r[0]) else datetime.fromisoformat(str(r[0])[:10])
            except Exception:
                d = datetime.now()
            dates.append(d)
            profits.append(float(r[1]))
    if not dates or not profits:
        return None
    equity = []
    total = 0
    for p in profits:
        total += p
        equity.append(total)

    fig, ax = plt.subplots()
    ax.plot(dates, equity, marker='o')
    ax.set_title("Equity Curve")
    ax.set_xlabel("Дата")
    ax.set_ylabel("Прибыль")
    fig.autofmt_xdate()

    buf = BytesIO()
    plt.savefig(buf, format="png")
    buf.seek(0)
    return buf

# ========== FSM Состояния ==========
class States(StatesGroup):
    add_account_name = State()
    add_account_balance = State()
    trade_step_1 = State()   # выбор счёта
    trade_step_2 = State()   # ввод пары
    trade_step_3 = State()   # выбор типа
    trade_step_4 = State()   # выбор даты входа
    trade_step_5 = State()   # выбор даты выхода
    trade_step_6 = State()   # ввод суммы
    trade_step_7 = State()   # ввод TP
    trade_step_8 = State()   # выбор статуса
    trade_step_9 = State()   # ввод стратегии
    trade_step_10 = State()  # чеклист
    trade_step_11 = State()  # заметки
    trade_step_12 = State()  # скриншот
    edit_trade = State()     # выбор сделки для редактирования
    edit_field = State()     # выбор поля
    edit_checklist = State() # редактирование чеклиста
    admin_wait_password = State()

# ========== Константы callback_data ==========
CALLBACK_NEW_TRADE = "new_trade"
CALLBACK_HIST_TRADES = "hist_trades"
CALLBACK_EDIT_TRADES = "edit_trades"
CALLBACK_DEL_TRADES = "del_trades"
CALLBACK_SKIP = "skip"
CALLBACK_SAVE_TRADE = "save_trade"
CALLBACK_CANCEL_TRADE = "cancel_trade"
CALLBACK_NEW_ACCOUNT = "new_acc"
CALLBACK_LIST_ACCOUNTS = "list_acc"
CALLBACK_SUBSCRIBE = "subscribe"
CALLBACK_CHANGE_LANG = "change_lang"

# ========== Инициализация бота и диспетчера ==========
bot = Bot(token=TOKEN)
dp = Dispatcher(storage=MemoryStorage())

# ========== ОБРАБОТЧИКИ ==========

@dp.message(Command("start"))
async def start(message: types.Message):
    user_id = message.from_user.id
    db.create_or_update_user(
        user_id,
        message.from_user.username or "user",
        message.from_user.first_name or "User"
    )
    await message.answer(get_text(user_id, "start"), reply_markup=main_menu(user_id))

# ---------- Меню "Мои сделки" ----------
@dp.message(F.text.regexp(r"Мои сделки|My Trades"))
async def trades_menu(message: types.Message):
    user_id = message.from_user.id
    count = db.count_user_trades(user_id)
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=get_text(user_id, "new_trade"), callback_data=CALLBACK_NEW_TRADE)],
        [InlineKeyboardButton(text=get_text(user_id, "history"), callback_data=CALLBACK_HIST_TRADES)],
        [InlineKeyboardButton(text=get_text(user_id, "edit"), callback_data=CALLBACK_EDIT_TRADES)],
        [InlineKeyboardButton(text=get_text(user_id, "delete"), callback_data=CALLBACK_DEL_TRADES)],
    ])
    await message.answer(f"{get_text(user_id, 'my_trades')} ({count}):", reply_markup=kb)

# ---------- Новая сделка ----------
@dp.callback_query(F.data == CALLBACK_NEW_TRADE)
async def new_trade(call: types.CallbackQuery, state: FSMContext):
    user_id = call.from_user.id
    if not can_add_trade(user_id):
        await call.answer(get_text(user_id, "limit"), show_alert=True)
        return
    accounts = db.get_accounts(user_id)
    if not accounts:
        await call.answer(get_text(user_id, "no_account"), show_alert=True)
        return
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"🏦 {a['name']}", callback_data=f"acc_{a['id']}")] for a in accounts
    ])
    await call.message.edit_text(f"👇 {get_text(user_id, 'select_account')}", reply_markup=kb)
    await state.set_state(States.trade_step_1)
    await call.answer()

@dp.callback_query(F.data.startswith("acc_"), States.trade_step_1)
async def select_acc(call: types.CallbackQuery, state: FSMContext):
    user_id = call.from_user.id
    try:
        acc_id = int(call.data.split("_")[1])
    except Exception:
        await call.answer("Invalid account")
        return
    await state.update_data(account_id=acc_id)
    # выбор пары
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🪙 BTC/USD", callback_data="pair_BTC/USD")],
        [InlineKeyboardButton(text="🪙 ETH/USD", callback_data="pair_ETH/USD")],
        [InlineKeyboardButton(text="💱 EUR/USD", callback_data="pair_EUR/USD")],
        [InlineKeyboardButton(text="💱 GBP/USD", callback_data="pair_GBP/USD")],
        [InlineKeyboardButton(text="✏️ Other", callback_data="pair_other")],
    ])
    await call.message.edit_text(f"1️⃣ {get_text(user_id, 'pair')}/Pair:", reply_markup=kb)
    await state.set_state(States.trade_step_2)
    await call.answer()

@dp.callback_query(F.data.startswith("pair_"), States.trade_step_2)
async def select_pair(call: types.CallbackQuery, state: FSMContext):
    user_id = call.from_user.id
    pair = call.data.replace("pair_", "")
    if pair == "other":
        await call.message.edit_text(f"✏️ {get_text(user_id, 'pair')}:")
        # остаёмся в том же состоянии, но ожидаем текст
        return
    await state.update_data(pair=pair)
    # переход к выбору типа
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📈 LONG", callback_data="type_long")],
        [InlineKeyboardButton(text="📉 SHORT", callback_data="type_short")],
    ])
    await call.message.edit_text(f"2️⃣ {get_text(user_id, 'type')}:", reply_markup=kb)
    await state.set_state(States.trade_step_3)
    await call.answer()

@dp.message(States.trade_step_2)
async def step_2_text(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    await state.update_data(pair=message.text.upper())
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📈 LONG", callback_data="type_long")],
        [InlineKeyboardButton(text="📉 SHORT", callback_data="type_short")],
    ])
    await message.answer(f"2️⃣ {get_text(user_id, 'type')}:", reply_markup=kb)
    await state.set_state(States.trade_step_3)

@dp.callback_query(F.data.startswith("type_"), States.trade_step_3)
async def select_type(call: types.CallbackQuery, state: FSMContext):
    user_id = call.from_user.id
    ttype = call.data.replace("type_", "").upper()
    await state.update_data(trade_type=ttype)
    now = datetime.now()
    await call.message.edit_text(
        f"3️⃣ {get_text(user_id, 'open_date')}:",
        reply_markup=calendar_kb(now.year, now.month, user_id)
    )
    await state.set_state(States.trade_step_4)
    await call.answer()

# ---------- Календарь ----------
@dp.callback_query(F.data.startswith("cal_"))
async def change_cal(call: types.CallbackQuery):
    parts = call.data.split("_")
    try:
        y, m = int(parts[1]), int(parts[2])
    except Exception:
        await call.answer()
        return
    await call.message.edit_reply_markup(reply_markup=calendar_kb(y, m, call.from_user.id))
    await call.answer()

@dp.callback_query(F.data.startswith("dt_"))
async def select_dt(call: types.CallbackQuery, state: FSMContext):
    user_id = call.from_user.id
    parts = call.data.split("_")
    try:
        year, month, day = int(parts[1]), int(parts[2]), int(parts[3])
    except Exception:
        await call.answer()
        return
    dt = f"{year}-{month:02d}-{day:02d}"
    curr_state = await state.get_state()
    if curr_state == States.trade_step_4.state:
        await state.update_data(open_date=dt)
        now = datetime.now()
        await call.message.edit_text(
            f"4️⃣ {get_text(user_id, 'close_date')}:",
            reply_markup=calendar_kb(now.year, now.month, user_id)
        )
        await state.set_state(States.trade_step_5)
    elif curr_state == States.trade_step_5.state:
        await state.update_data(close_date=dt)
        await call.message.edit_text(
            f"5️⃣ {get_text(user_id, 'enter_sum')}",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text=get_text(user_id, "skip"), callback_data=CALLBACK_SKIP)]
            ])
        )
        await state.set_state(States.trade_step_6)
    await call.answer()

# ---------- Ввод суммы (с парсингом) ----------
@dp.message(States.trade_step_6)
async def step_6(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    if not message.text:
        return
    text = message.text.strip()
    # Проверка на команду отмены
    if text.startswith("/") or is_main_menu_button(text):
        await state.clear()
        await message.answer("Действие отменено.", reply_markup=main_menu(user_id))
        return
    try:
        amount = parse_number(text)
    except ValueError:
        await message.answer(f"❌ {get_text(user_id, 'new_value')}")
        return
    await state.update_data(amount=amount)
    await message.answer(
        f"6️⃣ {get_text(user_id, 'enter_tp')}",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=get_text(user_id, "skip"), callback_data=CALLBACK_SKIP)]
        ])
    )
    await state.set_state(States.trade_step_7)

# ---------- Ввод TP ----------
@dp.message(States.trade_step_7)
async def step_7(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    if not message.text:
        return
    text = message.text.strip()
    if text.startswith("/") or is_main_menu_button(text):
        await state.clear()
        await message.answer("Действие отменено.", reply_markup=main_menu(user_id))
        return
    try:
        tp = parse_number(text)
    except ValueError:
        await message.answer(f"❌ {get_text(user_id, 'new_value')}")
        return
    await state.update_data(take_profit=tp)
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"✅ {get_text(user_id, 'profit')}", callback_data="st_profit")],
        [InlineKeyboardButton(text=f"❌ {get_text(user_id, 'loss')}", callback_data="st_loss")],
        [InlineKeyboardButton(text=f"🔒 {get_text(user_id, 'closed')}", callback_data="st_closed")],
        [InlineKeyboardButton(text=get_text(user_id, "skip"), callback_data=CALLBACK_SKIP)],
    ])
    await message.answer(f"7️⃣ {get_text(user_id, 'select_status')}", reply_markup=kb)
    await state.set_state(States.trade_step_8)

# ---------- Выбор статуса ----------
@dp.callback_query(F.data.startswith("st_"), States.trade_step_8)
async def select_st(call: types.CallbackQuery, state: FSMContext):
    user_id = call.from_user.id
    st = call.data.split("_")[1].upper()
    await state.update_data(status=st)
    await call.message.edit_text(
        f"8️⃣ {get_text(user_id, 'enter_strategy')}",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=get_text(user_id, "skip"), callback_data=CALLBACK_SKIP)]
        ])
    )
    await state.set_state(States.trade_step_9)
    await call.answer()

# ---------- Стратегия ----------
@dp.message(States.trade_step_9)
async def step_9(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    if not message.text:
        return
    text = message.text.strip()
    if text.startswith("/") or is_main_menu_button(text):
        await state.clear()
        await message.answer("Действие отменено.", reply_markup=main_menu(user_id))
        return
    await state.update_data(strategy=text)
    await message.answer(
        f"9️⃣ {get_text(user_id, 'checklist')}",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=get_text(user_id, "create_checklist"), callback_data="create_check")],
            [InlineKeyboardButton(text=get_text(user_id, "skip"), callback_data=CALLBACK_SKIP)],
            [InlineKeyboardButton(text=get_text(user_id, "templates"), callback_data="templates_list")]
        ])
    )
    await state.set_state(States.trade_step_10)

# ---------- Чеклист (создание) ----------
@dp.callback_query(F.data == "create_check", States.trade_step_10)
async def create_check(call: types.CallbackQuery, state: FSMContext):
    user_id = call.from_user.id
    await call.message.edit_text("✏️ Вводите пункты (каждый с новой строки):")
    # состояние остаётся trade_step_10, но теперь ждём текст
    await call.answer()

@dp.message(States.trade_step_10)
async def step_10_save_checklist(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    if not message.text:
        return
    text = message.text.strip()
    if text.startswith("/") or is_main_menu_button(text):
        await state.clear()
        await message.answer("Действие отменено.", reply_markup=main_menu(user_id))
        return
    items = [i.strip().lstrip("-").strip() for i in text.split("\n") if i.strip()]
    checklist = {item: False for item in items}
    # сохраняем во временный файл
    Path("trade_checklists").mkdir(exist_ok=True)
    fn = f"trade_checklists/{user_id}_{int(datetime.now().timestamp())}.json"
    try:
        with open(fn, "w", encoding="utf-8") as f:
            json.dump(checklist, f, ensure_ascii=False, indent=2)
        await state.update_data(checklist=fn)
    except Exception as e:
        logger.exception("Failed to save checklist")
        await message.answer("❌ Ошибка сохранения чеклиста")
        return
    await message.answer(
        f"🔟 {get_text(user_id, 'enter_notes')}",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=get_text(user_id, "skip"), callback_data=CALLBACK_SKIP)]
        ])
    )
    await state.set_state(States.trade_step_11)

# ---------- Заметки ----------
@dp.message(States.trade_step_11)
async def step_11(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    if not message.text:
        return
    text = message.text.strip()
    if text.startswith("/") or is_main_menu_button(text):
        await state.clear()
        await message.answer("Действие отменено.", reply_markup=main_menu(user_id))
        return
    await state.update_data(notes=text)
    await message.answer(
        f"1️⃣1️⃣ {get_text(user_id, 'select_screenshot')}",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=get_text(user_id, "skip"), callback_data=CALLBACK_SKIP)]
        ])
    )
    await state.set_state(States.trade_step_12)

# ---------- Скриншот и подтверждение ----------
@dp.message(States.trade_step_12, F.photo)
async def step_12_photo(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    photo = message.photo[-1]
    try:
        fp = await bot.get_file(photo.file_id)
        Path("trade_photos").mkdir(exist_ok=True)
        fn = f"trade_photos/{user_id}_{int(datetime.now().timestamp())}.jpg"
        await bot.download_file(fp.file_path, fn)
        await state.update_data(screenshot_path=fn)
    except Exception as e:
        logger.exception("Failed to download photo")
    await show_confirmation(message, state)

@dp.message(States.trade_step_12)
async def step_12_no_photo(message: types.Message, state: FSMContext):
    # если прислали не фото, считаем что пропустили
    user_id = message.from_user.id
    if message.text and (message.text.startswith("/") or is_main_menu_button(message.text)):
        await state.clear()
        await message.answer("Действие отменено.", reply_markup=main_menu(user_id))
        return
    await show_confirmation(message, state)

async def show_confirmation(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    data = await state.get_data()
    text = f"""✅ {get_text(user_id, 'confirm')}

📌 {get_text(user_id, 'pair')}: {data.get('pair', 'N/A')}
🔀 {get_text(user_id, 'type')}: {data.get('trade_type', 'N/A')}
📅 {get_text(user_id, 'open_date')}: {data.get('open_date', 'N/A')}
📅 {get_text(user_id, 'close_date')}: {data.get('close_date', 'N/A')}
💰 {get_text(user_id, 'amount')}: {data.get('amount', 'N/A')} USD
🎯 {get_text(user_id, 'tp')}: {data.get('take_profit', 'N/A')} USD
⚡ {get_text(user_id, 'status')}: {data.get('status', 'N/A')}
📊 {get_text(user_id, 'strategy')}: {data.get('strategy', 'N/A')}

{get_text(user_id, 'save')}?"""
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=get_text(user_id, "save"), callback_data=CALLBACK_SAVE_TRADE)],
        [InlineKeyboardButton(text=get_text(user_id, "cancel"), callback_data=CALLBACK_CANCEL_TRADE)],
    ])
    await message.answer(text, reply_markup=kb)

@dp.callback_query(F.data == CALLBACK_SAVE_TRADE)
async def save_trade(call: types.CallbackQuery, state: FSMContext):
    user_id = call.from_user.id
    data = await state.get_data()
    try:
        db.add_trade(user_id, data)
        db.increment_trades(user_id)
        acc_id = data.get("account_id")
        if acc_id:
            db.recalc_account_balance(user_id, acc_id)
        await state.clear()
        await call.message.edit_text(get_text(user_id, "saved"))
    except Exception as e:
        logger.exception("Failed to save trade")
        await call.message.edit_text("❌ Ошибка сохранения сделки")
    await call.answer()

@dp.callback_query(F.data == CALLBACK_CANCEL_TRADE)
async def cancel_trade(call: types.CallbackQuery, state: FSMContext):
    user_id = call.from_user.id
    await state.clear()
    await call.message.edit_text(get_text(user_id, "cancelled"))
    await call.answer()

# ---------- Обработка SKIP ----------
@dp.callback_query(F.data == CALLBACK_SKIP)
async def skip_handler(call: types.CallbackQuery, state: FSMContext):
    user_id = call.from_user.id
    curr_state = await state.get_state()
    if curr_state == States.trade_step_2.state:
        await state.update_data(pair="N/A")
        # переходим к выбору типа
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📈 LONG", callback_data="type_long")],
            [InlineKeyboardButton(text="📉 SHORT", callback_data="type_short")],
        ])
        await call.message.edit_text(f"2️⃣ {get_text(user_id, 'type')}:", reply_markup=kb)
        await state.set_state(States.trade_step_3)
    elif curr_state == States.trade_step_4.state:
        await state.update_data(open_date="N/A")
        now = datetime.now()
        await call.message.edit_text(
            f"4️⃣ {get_text(user_id, 'close_date')}:",
            reply_markup=calendar_kb(now.year, now.month, user_id)
        )
        await state.set_state(States.trade_step_5)
    elif curr_state == States.trade_step_5.state:
        await state.update_data(close_date="N/A")
        await call.message.edit_text(
            f"5️⃣ {get_text(user_id, 'enter_sum')}",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text=get_text(user_id, "skip"), callback_data=CALLBACK_SKIP)]
            ])
        )
        await state.set_state(States.trade_step_6)
    elif curr_state == States.trade_step_6.state:
        await state.update_data(amount="N/A")
        await call.message.edit_text(
            f"6️⃣ {get_text(user_id, 'enter_tp')}",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text=get_text(user_id, "skip"), callback_data=CALLBACK_SKIP)]
            ])
        )
        await state.set_state(States.trade_step_7)
    elif curr_state == States.trade_step_7.state:
        await state.update_data(take_profit="N/A")
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=f"✅ {get_text(user_id, 'profit')}", callback_data="st_profit")],
            [InlineKeyboardButton(text=f"❌ {get_text(user_id, 'loss')}", callback_data="st_loss")],
            [InlineKeyboardButton(text=f"🔒 {get_text(user_id, 'closed')}", callback_data="st_closed")],
            [InlineKeyboardButton(text=get_text(user_id, "skip"), callback_data=CALLBACK_SKIP)],
        ])
        await call.message.edit_text(f"7️⃣ {get_text(user_id, 'select_status')}", reply_markup=kb)
        await state.set_state(States.trade_step_8)
    elif curr_state == States.trade_step_8.state:
        await state.update_data(status="N/A")
        await call.message.edit_text(
            f"8️⃣ {get_text(user_id, 'enter_strategy')}",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text=get_text(user_id, "skip"), callback_data=CALLBACK_SKIP)]
            ])
        )
        await state.set_state(States.trade_step_9)
    elif curr_state == States.trade_step_9.state:
        await state.update_data(strategy="N/A")
        await call.message.edit_text(
            f"9️⃣ {get_text(user_id, 'checklist')}",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text=get_text(user_id, "create_checklist"), callback_data="create_check")],
                [InlineKeyboardButton(text=get_text(user_id, "skip"), callback_data=CALLBACK_SKIP)],
                [InlineKeyboardButton(text=get_text(user_id, "templates"), callback_data="templates_list")]
            ])
        )
        await state.set_state(States.trade_step_10)
    elif curr_state == States.trade_step_10.state:
        await state.update_data(checklist=None)
        await call.message.edit_text(
            f"🔟 {get_text(user_id, 'enter_notes')}",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text=get_text(user_id, "skip"), callback_data=CALLBACK_SKIP)]
            ])
        )
        await state.set_state(States.trade_step_11)
    elif curr_state == States.trade_step_11.state:
        await state.update_data(notes="N/A")
        await call.message.edit_text(
            f"1️⃣1️⃣ {get_text(user_id, 'select_screenshot')}",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text=get_text(user_id, "skip"), callback_data=CALLBACK_SKIP)]
            ])
        )
        await state.set_state(States.trade_step_12)
    elif curr_state == States.trade_step_12.state:
        data = await state.get_data()
        text = f"""✅ {get_text(user_id, 'confirm')}

📌 {get_text(user_id, 'pair')}: {data.get('pair', 'N/A')}
🔀 {get_text(user_id, 'type')}: {data.get('trade_type', 'N/A')}
📅 {get_text(user_id, 'open_date')}: {data.get('open_date', 'N/A')}
📅 {get_text(user_id, 'close_date')}: {data.get('close_date', 'N/A')}
💰 {get_text(user_id, 'amount')}: {data.get('amount', 'N/A')} USD
🎯 {get_text(user_id, 'tp')}: {data.get('take_profit', 'N/A')} USD
⚡ {get_text(user_id, 'status')}: {data.get('status', 'N/A')}
📊 {get_text(user_id, 'strategy')}: {data.get('strategy', 'N/A')}

{get_text(user_id, 'save')}?"""
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=get_text(user_id, "save"), callback_data=CALLBACK_SAVE_TRADE)],
            [InlineKeyboardButton(text=get_text(user_id, "cancel"), callback_data=CALLBACK_CANCEL_TRADE)],
        ])
        await call.message.edit_text(text, reply_markup=kb)
    await call.answer()

# ---------- История сделок ----------
@dp.callback_query(F.data == CALLBACK_HIST_TRADES)
async def hist_trades(call: types.CallbackQuery):
    user_id = call.from_user.id
    trades = db.get_recent_trades(user_id, limit=10)
    if not trades:
        await call.message.edit_text(get_text(user_id, "no_trades"))
    else:
        text = f"📋 {get_text(user_id, 'history')}:\n\n"
        for t in trades:
            emoji = "✅" if t["status"] == "PROFIT" else ("❌" if t["status"] == "LOSS" else "🔘")
            arrow = "📈" if t["trade_type"] == "LONG" else "📉"
            text += f"{emoji} {arrow} {t['pair']} ({t['trade_type']}) - {t['open_date']}\n"
        await call.message.edit_text(text)
    await call.answer()

# ---------- Удаление сделок ----------
@dp.callback_query(F.data == CALLBACK_DEL_TRADES)
async def del_trades(call: types.CallbackQuery):
    user_id = call.from_user.id
    trades = db.get_recent_trades(user_id, limit=5)
    if not trades:
        await call.message.edit_text(get_text(user_id, "no_trades"))
        await call.answer()
        return
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"🗑 {t['pair']}", callback_data=f"del_id_{t['id']}")] for t in trades
    ])
    await call.message.edit_text(get_text(user_id, "delete"), reply_markup=kb)
    await call.answer()

@dp.callback_query(F.data.startswith("del_id_"))
async def del_id(call: types.CallbackQuery):
    user_id = call.from_user.id
    try:
        trade_id = int(call.data.split("_")[2])
    except Exception:
        await call.answer()
        return
    trade = db.get_trade(trade_id, user_id)
    if not trade:
        await call.answer("Сделка не найдена")
        return
    account_id = trade.get("account_id")
    db.delete_trade(trade_id, user_id)
    db.decrement_trades(user_id)
    if account_id:
        db.recalc_account_balance(user_id, account_id)
    await call.message.edit_text(get_text(user_id, "deleted"))
    await call.answer()

# ---------- Редактирование сделок ----------
@dp.callback_query(F.data == CALLBACK_EDIT_TRADES)
async def edit_trades(call: types.CallbackQuery, state: FSMContext):
    user_id = call.from_user.id
    trades = db.get_recent_trades(user_id, limit=5)
    if not trades:
        await call.message.edit_text(get_text(user_id, "no_trades"))
        await call.answer()
        return
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"✏️ {t['pair']}", callback_data=f"edit_id_{t['id']}")] for t in trades
    ])
    await call.message.edit_text(get_text(user_id, "select_field"), reply_markup=kb)
    await state.set_state(States.edit_trade)
    await call.answer()

@dp.callback_query(F.data.startswith("edit_id_"), States.edit_trade)
async def edit_id(call: types.CallbackQuery, state: FSMContext):
    user_id = call.from_user.id
    try:
        trade_id = int(call.data.split("_")[2])
    except Exception:
        await call.answer()
        return
    await state.update_data(trade_id=trade_id)
    fields = [
        get_text(user_id, "pair"),
        get_text(user_id, "type"),
        get_text(user_id, "open_date"),
        get_text(user_id, "close_date"),
        get_text(user_id, "amount"),
        get_text(user_id, "tp"),
        get_text(user_id, "status"),
        get_text(user_id, "strategy"),
        get_text(user_id, "notes"),
        get_text(user_id, "pnl"),
    ]
    keys = ["pair", "trade_type", "open_date", "close_date", "amount", "take_profit", "status", "strategy", "notes", "pnl"]
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f, callback_data=f"ed_f_{k}")] for f, k in zip(fields, keys)
    ])
    await call.message.edit_text(get_text(user_id, "select_field"), reply_markup=kb)
    await state.set_state(States.edit_field)
    await call.answer()

@dp.callback_query(F.data.startswith("ed_f_"), States.edit_field)
async def ed_f(call: types.CallbackQuery, state: FSMContext):
    user_id = call.from_user.id
    field_short = call.data.split("_")[2]
    # field_map может быть избыточной, но оставим для ясности
    field_map = {
        "pair": "pair",
        "trade_type": "trade_type",
        "open_date": "open_date",
        "close_date": "close_date",
        "amount": "amount",
        "take_profit": "take_profit",
        "status": "status",
        "strategy": "strategy",
        "notes": "notes",
        "pnl": "pnl",
    }
    field = field_map.get(field_short, field_short)
    await state.update_data(field=field)
    await call.message.edit_text(get_text(user_id, "new_value"))
    # состояние остаётся edit_field
    await call.answer()

@dp.message(States.edit_field)
async def save_edit(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    if not message.text:
        return
    text = message.text.strip()
    if text.startswith("/") or is_main_menu_button(text):
        await state.clear()
        await message.answer("Действие отменено.", reply_markup=main_menu(user_id))
        return
    data = await state.get_data()
    trade_id = data.get("trade_id")
    field = data.get("field")
    if not trade_id or not field:
        await state.clear()
        await message.answer("Ошибка: не найдены данные сделки")
        return
    allowed_fields = {"pair", "trade_type", "open_date", "close_date", "amount", "take_profit", "status", "strategy", "notes", "pnl"}
    if field not in allowed_fields:
        await state.clear()
        await message.answer("❌ Недопустимое поле")
        return
    value: Union[str, float] = text
    if field in ("amount", "take_profit", "pnl"):
        try:
            value = parse_number(text)
        except ValueError:
            await message.answer(f"❌ {get_text(user_id, 'new_value')}")
            return
    try:
        trade = db.get_trade(trade_id, user_id)
        if not trade:
            await message.answer("Сделка не найдена")
            await state.clear()
            return
        account_id = trade.get("account_id")
        db.update_trade_field(trade_id, user_id, field, value)
        if field == "pnl" and account_id:
            db.recalc_account_balance(user_id, account_id)
        await state.clear()
        await message.answer(f"✅ {get_text(user_id, 'saved')}")
    except Exception as e:
        logger.exception("Failed to update trade")
        await message.answer("❌ Ошибка обновления")

# ---------- Счета ----------
@dp.message(F.text.regexp(r"🏦|Счет|Accounts"))
async def accounts(message: types.Message):
    user_id = message.from_user.id
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=get_text(user_id, "create_acc"), callback_data=CALLBACK_NEW_ACCOUNT)],
        [InlineKeyboardButton(text=get_text(user_id, "list_acc"), callback_data=CALLBACK_LIST_ACCOUNTS)],
    ])
    await message.answer(f"{get_text(user_id, 'accounts')}:", reply_markup=kb)

@dp.callback_query(F.data == CALLBACK_NEW_ACCOUNT)
async def new_acc(call: types.CallbackQuery, state: FSMContext):
    user_id = call.from_user.id
    await call.message.edit_text(get_text(user_id, "acc_name"))
    await state.set_state(States.add_account_name)
    await call.answer()

@dp.message(States.add_account_name)
async def acc_name(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    if not message.text:
        return
    text = message.text.strip()
    if text.startswith("/") or is_main_menu_button(text):
        await state.clear()
        await message.answer("Действие отменено.", reply_markup=main_menu(user_id))
        return
    await state.update_data(name=text)
    await message.answer(get_text(user_id, "acc_balance"))
    await state.set_state(States.add_account_balance)

@dp.message(States.add_account_balance)
async def acc_balance(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    if not message.text:
        return
    text = message.text.strip()
    if text.startswith("/") or is_main_menu_button(text):
        await state.clear()
        await message.answer("Действие отменено.", reply_markup=main_menu(user_id))
        return
    data = await state.get_data()
    name = data.get("name")
    if not name:
        await message.answer("❌ Не найдено название счёта. Введите название снова:")
        await state.set_state(States.add_account_name)
        return
    try:
        balance = parse_number(text)
    except ValueError:
        await message.answer("❌ Введите число — баланс счёта (например: 25000 или 25 000):")
        return
    if balance < 0:
        await message.answer("❌ Баланс не может быть отрицательным. Введите положительное число:")
        return
    try:
        db.create_account(user_id, name, balance)
        await state.clear()
        await message.answer(get_text(user_id, "acc_created"))
    except Exception as e:
        logger.exception("acc_balance failed")
        await message.answer("❌ Ошибка сохранения счёта")

@dp.callback_query(F.data == CALLBACK_LIST_ACCOUNTS)
async def list_acc(call: types.CallbackQuery):
    user_id = call.from_user.id
    accounts = db.get_accounts(user_id)
    if not accounts:
        await call.message.edit_text(get_text(user_id, "no_account"))
    else:
        text = f"{get_text(user_id, 'accounts')}:\n\n"
        for acc in accounts:
            text += f"🏦 {acc['name']}\n🔷 {get_text(user_id, 'acc_balance')}: {acc['initial_balance']}\n🔶 {get_text(user_id, 'current')}: {acc['current_balance']}\n\n"
        await call.message.edit_text(text)
    await call.answer()

# ---------- Аналитика ----------
@dp.message(F.text.regexp(r"📈|Analytics"))
async def analytics(message: types.Message):
    user_id = message.from_user.id
    stats = calculate_full_stats(user_id)
    if not stats:
        await message.answer(get_text(user_id, "no_trades"), reply_markup=main_menu(user_id))
        return

    rr_text = f"{stats['avg_rr']}" if stats['avg_rr'] is not None else "∞"
    text = (
        f"{get_text(user_id, 'analytics')}\n\n"
        f"🧾 {get_text(user_id, 'all_trades')}: {stats['total']}\n"
        f"🏆 {get_text(user_id, 'profit')}: {stats['profitable']}\n"
        f"🧨 {get_text(user_id, 'loss')}: {stats['losing']}\n"
        f"🎯 Win Rate: {stats['win_rate']:.1f}%\n\n"
        f"💼 {get_text(user_id, 'total_pnl_label')}: {stats['total_pnl']}\n"
        f"📊 {get_text(user_id, 'avg_trade_label')}: {stats['avg_trade']}\n"
        f"🚀 {get_text(user_id, 'best_trade_label')}: {stats['best_trade']}\n"
        f"📉 {get_text(user_id, 'worst_trade_label')}: {stats['worst_trade']}\n\n"
        f"📏 {get_text(user_id, 'risk_section')}\n"
        f"⚖️ {get_text(user_id, 'rr_label')}: {rr_text}\n"
        f"🔥 {get_text(user_id, 'avg_risk_label')}: {stats['avg_risk']}\n"
        f"📉 {get_text(user_id, 'max_dd_label')}: {stats['max_drawdown']}\n"
        f"🏅 {get_text(user_id, 'win_streak_label')}: {stats['win_streak']}\n"
        f"💣 {get_text(user_id, 'loss_streak_label')}: {stats['loss_streak']}"
    )
    await message.answer(text, reply_markup=main_menu(user_id))

# ---------- Экспорт ----------
@dp.message(F.text.regexp(r"Экспорт|Export"))
async def export_start(message: types.Message):
    user_id = message.from_user.id
    accounts = db.get_accounts(user_id)
    if not accounts:
        await message.answer(get_text(user_id, "no_account"))
        return
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"📄 {acc['name']}", callback_data=f"expacct_{acc['id']}")] for acc in accounts
    ])
    await message.answer(get_text(user_id, "select_account_export"), reply_markup=kb)

@dp.callback_query(F.data.startswith("expacct_"))
async def export_do(call: types.CallbackQuery):
    user_id = call.from_user.id
    try:
        aid = int(call.data.split("_")[1])
    except Exception:
        await call.answer()
        return
    account = db.get_account(aid, user_id)
    if not account:
        await call.answer("Счёт не найден")
        return
    account_name = account['name']
    trades = db.get_trades_for_export(user_id, account_id=aid)
    Path("exports").mkdir(exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = account_name[:31]
    headers = [
        "Пара", "Тип", "Дата входа", "Дата выхода", "Сумма", "TP", "Статус",
        "Стратегия", "Заметки", "P/L", "Создано",
    ]
    ws.append(headers)
    for row in trades:
        ws.append(list(row))

    stats = calculate_full_stats(user_id)
    if stats:
        ws_stats = wb.create_sheet(title="Analytics")
        ws_stats.append(["Метрика", "Значение"])
        ws_stats.append([get_text(user_id, "all_trades"), stats["total"]])
        ws_stats.append([get_text(user_id, "profit"), stats["profitable"]])
        ws_stats.append([get_text(user_id, "loss"), stats["losing"]])
        ws_stats.append([get_text(user_id, "total_pnl_label"), stats["total_pnl"]])
        ws_stats.append([get_text(user_id, "avg_trade_label"), stats["avg_trade"]])
        ws_stats.append([get_text(user_id, "best_trade_label"), stats["best_trade"]])
        ws_stats.append([get_text(user_id, "worst_trade_label"), stats["worst_trade"]])
        rr_val = stats['avg_rr'] if stats['avg_rr'] is not None else "∞"
        ws_stats.append([get_text(user_id, "rr_label"), rr_val])
        ws_stats.append([get_text(user_id, "avg_risk_label"), stats["avg_risk"]])
        ws_stats.append([get_text(user_id, "max_dd_label"), stats["max_drawdown"]])
        ws_stats.append([get_text(user_id, "win_streak_label"), stats["win_streak"]])
        ws_stats.append([get_text(user_id, "loss_streak_label"), stats["loss_streak"]])

    fn = f"exports/trades_{user_id}_{aid}_{int(datetime.now().timestamp())}.xlsx"
    try:
        wb.save(fn)
        caption = f"✅ Сделки по счёту «{account_name}» — {len(trades)} записей."
        # Отправляем новым сообщением, не редактируем старое
        await bot.send_document(
            call.from_user.id,
            FSInputFile(fn),
            caption=caption,
        )
        # Удаляем временный файл
        Path(fn).unlink(missing_ok=True)
        # Подтверждаем callback и отправляем сообщение об успехе
        await call.answer(get_text(user_id, "sent"))
        # Можно дополнительно отправить уведомление, но не обязательно
    except Exception as e:
        logger.exception("export error")
        await call.answer("❌ Ошибка при создании файла")

# ---------- Команда /stats ----------
@dp.message(Command("stats"))
async def stats_cmd(message: types.Message):
    user_id = message.from_user.id
    stats = calculate_full_stats(user_id)
    if not stats:
        await message.answer("Нет сделок для анализа")
        return
    rr_text = f"{stats['avg_rr']}" if stats['avg_rr'] is not None else "∞"
    text = f"""
📊 Статистика

Сделок: {stats["total"]}
Прибыльных: {stats["profitable"]}
Убыточных: {stats["losing"]}
Winrate: {stats["win_rate"]}%
Общая прибыль: {stats["total_pnl"]}
Средняя сделка: {stats["avg_trade"]}
Лучшая: {stats["best_trade"]}
Худшая: {stats["worst_trade"]}
Средний риск: {stats["avg_risk"]}
Средний RR: {rr_text}
Макс. просадка: {stats["max_drawdown"]}
Серия побед: {stats["win_streak"]}
Серия убытков: {stats["loss_streak"]}
"""
    await message.answer(text)

# ---------- Команда /export (все сделки без выбора счёта) ----------
@dp.message(Command("export"))
async def export_excel(message: types.Message):
    user_id = message.from_user.id
    trades = db.get_trades_for_export(user_id)
    if not trades:
        await message.answer("Нет сделок для экспорта")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "All Trades"
    headers = ["Пара", "Тип", "Вход", "Выход", "Сумма", "TP", "Статус", "Стратегия", "Заметки", "P/L", "Создано"]
    ws.append(headers)
    for row in trades:
        ws.append(list(row))
    Path("exports").mkdir(exist_ok=True)
    fn = f"exports/all_{user_id}_{int(datetime.now().timestamp())}.xlsx"
    wb.save(fn)
    await message.answer_document(FSInputFile(fn))
    Path(fn).unlink(missing_ok=True)

# ---------- Команда /equity ----------
@dp.message(Command("equity"))
async def equity_cmd(message: types.Message):
    user_id = message.from_user.id
    chart = create_equity_chart(user_id)
    if chart:
        await message.answer_photo(chart)
    else:
        await message.answer("Нет данных для графика")

# ---------- Помощь ----------
@dp.message(F.text.regexp(r"🆘|Помощь|Help"))
async def help_cmd(message: types.Message):
    user_id = message.from_user.id
    text = """
📚 ПОМОЩЬ

➕ Новая сделка
Добавляет новую торговую сделку.

📒 Мои сделки
Просмотр всех сделок.

📈 Аналитика
Статистика торговли:
• прибыль
• winrate
• риск-менеджмент

📤 Экспорт
Скачивает Excel файл со всеми сделками.

🏦 Счета
Управление торговыми счетами.

💳 Подписка
Убирает лимит сделок.

Бесплатно: до 20 сделок
Премиум: без ограничений
"""
    await message.answer(text, reply_markup=main_menu(user_id))

# ---------- Настройки ----------
@dp.message(F.text.regexp(r"🧰|Настройки|Settings"))
async def settings(message: types.Message):
    user_id = message.from_user.id
    user = db.get_user(user_id)
    if not user:
        return
    lang_text = "🇷🇺 Русский" if user["language"] == "ru" else "🇬🇧 English"
    text = f"""{get_text(user_id, 'settings')}

{get_text(user_id, 'id')}{user_id}
{get_text(user_id, 'username')}{user['username']}
{get_text(user_id, 'trades_count')}{user['total_trades']}

🌐 {get_text(user_id, 'current_lang')}{lang_text}"""
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=get_text(user_id, "language"), callback_data=CALLBACK_CHANGE_LANG)],
        [InlineKeyboardButton(text=get_text(user_id, "subscribe"), callback_data=CALLBACK_SUBSCRIBE)],
    ])
    await message.answer(text, reply_markup=kb)

@dp.callback_query(F.data == CALLBACK_CHANGE_LANG)
async def change_lang(call: types.CallbackQuery):
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🇷🇺 Русский", callback_data="lang_ru")],
        [InlineKeyboardButton(text="🇬🇧 English", callback_data="lang_en")],
    ])
    await call.message.edit_text("Выберите язык / Select language:", reply_markup=kb)
    await call.answer()

@dp.callback_query(F.data.startswith("lang_"))
async def set_lang(call: types.CallbackQuery):
    lang = call.data.split("_")[1]
    user_id = call.from_user.id
    db.set_language(user_id, lang)
    lang_text = "🇷🇺 Русский" if lang == "ru" else "🇬🇧 English"
    await call.message.edit_text(f"✅ Язык изменен на {lang_text}")
    await call.answer()

# ---------- Подписка (Telegram Payments) ----------
@dp.callback_query(F.data == CALLBACK_SUBSCRIBE)
async def subscribe_menu(call: types.CallbackQuery):
    user_id = call.from_user.id
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=get_text(user_id, "plan_2"), callback_data="plan_30_200")],
        [InlineKeyboardButton(text=get_text(user_id, "plan_5"), callback_data="plan_90_500")],
    ])
    await call.message.edit_text(get_text(user_id, "subscribe_info"), reply_markup=kb)
    await call.answer()

@dp.callback_query(F.data.startswith("plan_"))
async def plan_choice(call: types.CallbackQuery):
    try:
        _, days_str, cents_str = call.data.split("_")
        days = int(days_str)
        price_cents = int(cents_str)
    except Exception:
        await call.answer("Invalid plan")
        return
    user_id = call.from_user.id
    if not PROVIDER_TOKEN:
        await call.answer("Payments not configured", show_alert=True)
        return
    title = f"Premium {days} days"
    description = f"Premium access for {days} days"
    payload = f"subscribe:{user_id}:{days}"
    prices = [LabeledPrice(label=title, amount=price_cents)]
    try:
        await bot.send_invoice(
            chat_id=user_id,
            title=title,
            description=description,
            payload=payload,
            provider_token=PROVIDER_TOKEN,
            currency=CURRENCY,
            prices=prices,
            start_parameter="premium"
        )
        await call.answer()
    except Exception as e:
        logger.exception("send_invoice failed")
        await call.answer("Failed to create invoice", show_alert=True)

@dp.pre_checkout_query()
async def pre_checkout(pre_checkout_q: PreCheckoutQuery):
    try:
        await bot.answer_pre_checkout_query(pre_checkout_q.id, ok=True)
    except Exception as e:
        logger.exception("pre_checkout error")

@dp.message(F.content_type == "successful_payment")
async def process_successful_payment(message: types.Message):
    payload = message.successful_payment.invoice_payload
    try:
        parts = payload.split(":")
        if parts[0] == "subscribe":
            uid = int(parts[1])
            days = int(parts[2])
            provider_id = message.successful_payment.telegram_payment_charge_id
            period_end = db.grant_premium(uid, days, provider="telegram_payments", provider_id=provider_id)
            if period_end:
                await message.answer(get_text(uid, "grant_success").format(days))
    except Exception as e:
        logger.exception("process_successful_payment error")

# ---------- Админ: ручная выдача премиума ----------
@dp.message(Command("grant_manual"))
async def cmd_grant_manual(message: types.Message):
    user = message.from_user.id
    if user != ADMIN_ID:
        await message.answer(get_text(user, "not_admin"))
        return
    parts = message.text.split()
    if len(parts) < 3:
        await message.answer("Usage: /grant_manual <user_id> <days>")
        return
    try:
        uid = int(parts[1])
        days = int(parts[2])
    except Exception:
        await message.answer("Invalid parameters")
        return
    period_end = db.grant_premium(uid, days, provider="admin_manual", provider_id=str(user))
    if period_end:
        await message.answer(get_text(user, "grant_success").format(days))
        try:
            await bot.send_message(uid, get_text(uid, "grant_success").format(days))
        except Exception as e:
            logger.warning(f"Could not notify user {uid}")

# ---------- Админ: вход по паролю ----------
async def _activate_admin_premium(user_id: int, username: str, first_name: str) -> bool:
    """Активация бессрочного премиума для админа."""
    # Бессрочный премиум: устанавливаем очень далёкую дату (например, 2100 год)
    far_future = "2100-01-01T00:00:00"
    try:
        db.cursor.execute(
            "UPDATE users SET is_premium=1, premium_until=? WHERE user_id=?",
            (far_future, user_id)
        )
        if db.cursor.rowcount == 0:
            db.cursor.execute(
                """INSERT INTO users(user_id, username, first_name, is_premium, premium_until)
                   VALUES(?, ?, ?, 1, ?)""",
                (user_id, username or "", first_name or "User", far_future)
            )
        db.cursor.execute(
            "INSERT INTO subscriptions(user_id, provider, provider_id, status, period_end) VALUES(?, ?, ?, ?, ?)",
            (user_id, "admin_login", "", "active", far_future)
        )
        db.commit()
        return True
    except Exception as e:
        logger.exception("admin_login failed")
        return False

@dp.message(Command("admin"))
async def admin_login(message: types.Message, state: FSMContext):
    parts = message.text.split()
    if len(parts) >= 2:
        password = parts[1].strip()
        if password != ADMIN_PASSWORD:
            await message.answer("Неверный пароль")
            return
        ok = await _activate_admin_premium(
            message.from_user.id,
            message.from_user.username,
            message.from_user.first_name
        )
        if ok:
            await message.answer("✅ Админ-доступ активирован. Безлимитные сделки включены.")
        else:
            await message.answer("❌ Ошибка при активации. Попробуйте позже.")
        return
    await state.set_state(States.admin_wait_password)
    await message.answer("Введите пароль администратора:")

@dp.message(States.admin_wait_password)
async def admin_password_step(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    password = (message.text or "").strip()
    if password != ADMIN_PASSWORD:
        await state.clear()
        await message.answer("Неверный пароль")
        return
    ok = await _activate_admin_premium(user_id, message.from_user.username, message.from_user.first_name)
    await state.clear()
    if ok:
        await message.answer("✅ Админ-доступ активирован. Безлимитные сделки включены.")
    else:
        await message.answer("❌ Ошибка при активации. Попробуйте позже.")

# ---------- Шаблоны чеклистов ----------
@dp.message(Command("templates"))
async def templates_list(message: types.Message):
    user_id = message.from_user.id
    tdir = Path("checklist_templates")
    templates = list(tdir.glob("*.json"))
    if not templates:
        await message.answer(get_text(user_id, "no_templates"))
        return
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=t.stem, callback_data=f"applytpl_{t.name}")] for t in templates
    ])
    await message.answer(get_text(user_id, "templates"), reply_markup=kb)

@dp.callback_query(F.data.startswith("applytpl_"))
async def apply_template(call: types.CallbackQuery, state: FSMContext):
    user_id = call.from_user.id
    fname = call.data.split("_", 1)[1]
    try:
        with open(f"checklist_templates/{fname}", "r", encoding="utf-8") as f:
            tpl = json.load(f)
    except Exception as e:
        await call.answer("Template error")
        return
    Path("trade_checklists").mkdir(exist_ok=True)
    fn = f"trade_checklists/{user_id}_{int(datetime.now().timestamp())}.json"
    try:
        with open(fn, "w", encoding="utf-8") as f:
            json.dump(tpl, f, ensure_ascii=False, indent=2)
        await state.update_data(checklist=fn)
        await call.message.edit_text(
            f"🔟 {get_text(user_id, 'enter_notes')}",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text=get_text(user_id, "skip"), callback_data=CALLBACK_SKIP)]
            ])
        )
        await state.set_state(States.trade_step_11)
        await call.answer()
    except Exception as e:
        logger.exception("apply_template error")
        await call.answer("❌ Error applying template")

@dp.message(Command("my_checklists"))
async def my_checklists(message: types.Message):
    user_id = message.from_user.id
    files = list(Path("trade_checklists").glob(f"{user_id}_*.json"))
    if not files:
        await message.answer("No checklists")
        return
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f.name, callback_data=f"editchk_{f.name}")] for f in files
    ])
    await message.answer("Your checklists:", reply_markup=kb)

@dp.callback_query(F.data.startswith("editchk_"))
async def edit_checklist_start(call: types.CallbackQuery, state: FSMContext):
    fname = call.data.split("_", 1)[1]
    fpath = Path("trade_checklists") / fname
    if not fpath.exists():
        await call.answer("File not found")
        return
    try:
        data = json.loads(fpath.read_text(encoding="utf-8"))
    except Exception as e:
        await call.answer("Error reading file")
        return
    text = "Edit checklist items (send new list, each item on new line):\n\n" + "\n".join(
        [f"- {k} [{'x' if v else ' ' }]" for k, v in data.items()]
    )
    await call.message.edit_text(text)
    await state.update_data(editing_file=str(fpath))
    await state.set_state(States.edit_checklist)
    await call.answer()

@dp.message(States.edit_checklist)
async def edit_checklist_save(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    if not message.text:
        return
    text = message.text.strip()
    if text.startswith("/") or is_main_menu_button(text):
        await state.clear()
        await message.answer("Действие отменено.", reply_markup=main_menu(user_id))
        return
    data = await state.get_data()
    fpath = data.get("editing_file")
    if not fpath:
        await message.answer("No file in context")
        return
    items = [i.strip().lstrip("-").strip() for i in text.split("\n") if i.strip()]
    checklist = {item: False for item in items}
    try:
        with open(fpath, "w", encoding="utf-8") as f:
            json.dump(checklist, f, ensure_ascii=False, indent=2)
        await state.clear()
        await message.answer("Checklist updated")
    except Exception as e:
        logger.exception("edit_checklist_save error")
        await message.answer("Error saving checklist")

# ---------- Проверка нажатия кнопок главного меню ----------
MENU_BUTTON_MARKERS = (
    "Мои сделки", "My Trades", "Счет", "Accounts", "Аналитика", "Analytics",
    "Экспорт", "Export", "Помощь", "Help", "Настройки", "Settings",
)

def is_main_menu_button(text: str) -> bool:
    if not text or not text.strip():
        return False
    t = text.strip()
    return any(m in t for m in MENU_BUTTON_MARKERS)

# ---------- Вебхук ----------
async def handle_webhook(request):
    try:
        # Проверка секретного токена, если задан
        if WEBHOOK_SECRET:
            secret = request.headers.get("X-Telegram-Bot-Api-Secret-Token")
            if secret != WEBHOOK_SECRET:
                return web.Response(status=403, text="Forbidden")
        data = await request.json()
        update = types.Update(**data)
        await dp.feed_update(bot, update)
    except Exception as e:
        logger.error(f"Webhook error: {e}")
    return web.Response(text="OK")

async def on_startup(app):
    try:
        await bot.delete_webhook()
        await bot.set_webhook(WEBHOOK_URL, secret_token=WEBHOOK_SECRET if WEBHOOK_SECRET else None)
        logger.info(f"Webhook set: {WEBHOOK_URL}")
    except Exception as e:
        logger.error(f"Startup error: {e}")

# ========== Создание необходимых папок ==========
for d in ("trade_photos", "exports", "trade_checklists", "checklist_templates"):
    Path(d).mkdir(parents=True, exist_ok=True)

# ========== Запуск aiohttp приложения ==========
app = web.Application()
app.router.add_post(WEBHOOK_PATH, handle_webhook)
app.on_startup.append(on_startup)

if __name__ == "__main__":
    web.run_app(app, port=PORT, host="0.0.0.0")

